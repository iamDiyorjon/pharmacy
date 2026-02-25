"""
Drug import service — reads an Excel file and upserts medicines + availability.

Excel structure (adika.xlsx, sheet "TABE", data from row 7):
  B: Drug name (Russian)
  C: Manufacturer
  D: Expiry date (datetime)
  E: Price (UZS, integer)
  F: Quantity (float)
"""

from __future__ import annotations

import logging
from datetime import date, datetime
from pathlib import Path
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import select, and_
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.session import async_session
from app.models.medicine import Medicine, MedicineAvailability

logger = logging.getLogger(__name__)

DATA_START_ROW = 7
SHEET_NAME = "TABE"


def _parse_expiry(value) -> date | None:
    """Parse expiry date from Excel cell value."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _parse_number(value, default=None):
    """Parse a numeric cell value."""
    if value is None:
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


async def import_drugs_from_excel(file_path: str, pharmacy_id: UUID) -> dict:
    """Import drugs from an Excel file into the database.

    Returns a dict with stats: {new, updated, skipped, errors}.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")

    ws = wb[SHEET_NAME]

    stats = {"new": 0, "updated": 0, "skipped": 0, "errors": 0}

    async with async_session() as session:
        for row_num, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW), start=DATA_START_ROW):
            try:
                name_ru = row[1].value  # Column B (index 1)
                if not name_ru or not str(name_ru).strip():
                    stats["skipped"] += 1
                    continue

                name_ru = str(name_ru).strip()
                manufacturer = str(row[2].value).strip() if row[2].value else None  # Column C
                expiry_date = _parse_expiry(row[3].value)  # Column D
                price = _parse_number(row[4].value)  # Column E
                quantity = _parse_number(row[5].value, default=0)  # Column F

                # Upsert Medicine by name_ru
                result = await session.execute(
                    select(Medicine).where(Medicine.name_ru == name_ru)
                )
                medicine = result.scalar_one_or_none()

                if medicine is None:
                    medicine = Medicine(
                        name=name_ru,  # Use Russian name as primary name too
                        name_ru=name_ru,
                        manufacturer=manufacturer,
                    )
                    session.add(medicine)
                    await session.flush()
                    stats["new"] += 1
                else:
                    if manufacturer and medicine.manufacturer != manufacturer:
                        medicine.manufacturer = manufacturer
                    stats["updated"] += 1

                # Upsert MedicineAvailability
                result = await session.execute(
                    select(MedicineAvailability).where(
                        and_(
                            MedicineAvailability.medicine_id == medicine.id,
                            MedicineAvailability.pharmacy_id == pharmacy_id,
                        )
                    )
                )
                avail = result.scalar_one_or_none()

                if avail is None:
                    avail = MedicineAvailability(
                        medicine_id=medicine.id,
                        pharmacy_id=pharmacy_id,
                        is_available=quantity is not None and quantity > 0,
                        price=price,
                        quantity=quantity,
                        expiry_date=expiry_date,
                    )
                    session.add(avail)
                else:
                    avail.price = price
                    avail.quantity = quantity
                    avail.is_available = quantity is not None and quantity > 0
                    avail.expiry_date = expiry_date

            except Exception:
                logger.exception("Error processing row %d", row_num)
                stats["errors"] += 1
                continue

        await session.commit()

    wb.close()

    logger.info(
        "Import complete: %d new, %d updated, %d skipped, %d errors",
        stats["new"], stats["updated"], stats["skipped"], stats["errors"],
    )
    return stats
